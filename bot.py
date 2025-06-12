# -*- coding: utf-8 -*-
import telebot
import sqlite3
import os
from datetime import datetime
import pandas as pd
import json
import base64
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from telebot import types

# --- Настройки бота ---
BOT_TOKEN = '6425282545:AAHv28Q5sWLgMpMnn-9FLrASUYITQNMDFkM'
bot = telebot.TeleBot(BOT_TOKEN)

# --- ID менеджеров ---
MANAGER_IDS = [919034275, 372145026, 6432717873]

# --- Курсы валют ---
currency_rates = {
    "CNY": 14.5,
    "JPY": 0.72,
    "USD": 98.1,
    "EUR": 105.3
}

# --- Пути к файлам ---
script_dir = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(script_dir, 'volt_logistics.db')
orders_dir = os.path.join(script_dir, 'orders')
logs_dir = os.path.join(script_dir, 'logs')

os.makedirs(orders_dir, exist_ok=True)
os.makedirs(logs_dir, exist_ok=True)

log_file_path = os.path.join(logs_dir, f"volt_log_{datetime.now().strftime('%Y-%m-%d')}.txt")

# --- Логирование ---
def log_event(event_type, user_id, message=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file_path, "a", encoding="utf-8") as f:
        line = f"[{timestamp}] [{event_type}] [User {user_id}]"
        if message:
            line += f": {message}"
        line += "\n"
        f.write(line)
    print(f"[LOG] {event_type} | User: {user_id} | Message: {message}")

# --- Инициализация базы данных ---
def init_db():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    # Таблица пользователей
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            telegram_id INTEGER PRIMARY KEY,
            full_name TEXT,
            telegram_username TEXT,
            is_manager BOOLEAN DEFAULT FALSE
        )
    ''')
    # Таблица заказов
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            full_name TEXT,
            telegram_username TEXT,
            country TEXT,
            status TEXT DEFAULT 'На модерации',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # Таблица товаров внутри заказа
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS order_items (
            item_id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            photo BLOB,
            link_or_id TEXT,
            color TEXT,
            size TEXT,
            price REAL,
            quantity INTEGER,
            comment TEXT,
            FOREIGN KEY(order_id) REFERENCES orders(id)
        )
    ''')
    conn.commit()
    conn.close()
    log_event("init_db", "system", "База данных инициализирована.")

init_db()

# --- Добавление пользователя ---
def add_user(user_id, full_name, username):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT OR IGNORE INTO users (telegram_id, full_name, telegram_username) VALUES (?, ?, ?)",
        (user_id, full_name, username)
    )
    conn.commit()
    conn.close()
    log_event("user_registered", user_id, f"{full_name}, @{username}")

# --- Создание заказа ---
def create_order(user_id, full_name, telegram_username, country):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO orders 
        (user_id, full_name, telegram_username, country)
        VALUES (?, ?, ?, ?)
    ''', (user_id, full_name, telegram_username, country))
    order_id = cursor.lastrowid
    conn.commit()
    conn.close()
    log_event("order_created", user_id, f"Заказ #{order_id}, страна: {country}")
    return order_id

# --- Сохранение товара в заказе ---
def add_item_to_order(order_id, photo, link_or_id, color, size, price, quantity, comment=''):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO order_items 
        (order_id, photo, link_or_id, color, size, price, quantity, comment)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (order_id, photo, link_or_id, color, size, price, quantity, comment))
    conn.commit()
    conn.close()
    log_event("item_added", order_id,
              f"Фото: {photo}, Ссылка: {link_or_id}, Цена: {price}, Кол-во: {quantity}")

# --- Получение всех товаров для заказа ---
def get_items(order_id):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM order_items WHERE order_id=?", (order_id,))
    result = cursor.fetchall()
    conn.close()
    return result

# --- Путь к Excel ---
def get_daily_excel_path():
    today = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(orders_dir, f"{today}.xlsx")

# --- Вставка изображения в Excel ---
def insert_image_in_excel(file_path, image_base64, row_index):
    wb = load_workbook(file_path)
    ws = wb.active

    try:
        img_data = base64.b64decode(image_base64.split(",")[1])
        img = XLImage(BytesIO(img_data))
        ws.row_dimensions[row_index].height = 60
        ws.column_dimensions['A'].width = 20
        ws.add_image(img, f"A{row_index}")
    except Exception as e:
        log_event("excel_image_insert_failed", "system", str(e))

    wb.save(file_path)
    log_event("excel_image_inserted", "system", f"Фото добавлено в строку {row_index}")

# --- Сохранение в Excel с изображением ---
def save_to_excel(order_id, items, full_name, telegram_username, country, status, requisites, comment):
    daily_excel_path = get_daily_excel_path()

    try:
        df_existing = pd.read_excel(daily_excel_path)
    except Exception:
        df_existing = pd.DataFrame(columns=[
            'Номер заказа', 'Фото', 'ID / Ссылка', 'Цвет', 'Размер',
            'Цена за единицу', 'Кол-во/шт', 'tg контакт',
            'Страна выкупа', 'статус', 'реквизиты', 'комментарий'
        ])

    rows = []
    for idx, item in enumerate(items):
        rows.append({
            'Номер заказа': order_id,
            'Фото': item[1],  # base64 строка
            'ID / Ссылка': item[2],
            'Цвет': item[3],
            'Размер': item[4],
            'Цена за единицу': round(item[5], 2),
            'Кол-во/шт': item[6],
            'tg контакт': telegram_username,
            'Страна выкупа': country.upper(),
            'статус': status,
            'реквизиты': requisites,
            'комментарий': comment
        })

    df_new = pd.DataFrame(rows)
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    df_combined.to_excel(daily_excel_path, index=False)

    # Добавляем фото как изображения
    wb = load_workbook(daily_excel_path)
    ws = wb.active
    start_row = ws.max_row + 1
    for idx, item in enumerate(items):
        if item[1] and item[1].startswith("data:image"):
            insert_image_in_excel(daily_excel_path, item[1], start_row + idx)

    wb.save(daily_excel_path)
    log_event("excel_saved", "system", f"Данные записаны в {daily_excel_path}")

# --- Обработчик web_app_data ---
@bot.message_handler(content_types=['web_app_data'])
def handle_web_app_data(message):
    try:
        data = json.loads(message.web_app_data.data)
        user_id = message.from_user.id
        log_event("form_received", user_id, f"Получены данные: {data}")

        if not is_registered(user_id):
            bot.send_message(message.chat.id, "⚠️ Вы не зарегистрированы. Нажмите /start.")
            return

        if data.get('command') == 'create_order':
            full_name = data.get('full_name')
            telegram_username = data.get('telegram_username')
            country = data.get('country')
            items = data.get('items', [])

            if not all([full_name, telegram_username, country, len(items) > 0]):
                bot.send_message(message.chat.id, "⚠️ Не все поля формы заполнены.")
                return

            order_id = create_order(user_id, full_name, telegram_username, country)

            for item in items:
                photo_list = item.get('photos', [])
                photo = ', '.join(photo_list) if isinstance(photo_list, list) else ''
                link_or_id = item.get('link_or_id', '')
                color = item.get('color', '')
                size = item.get('size', '')
                price = float(item.get('price', 0))
                quantity = int(item.get('quantity', 1))
                comment = item.get('comment', '')

                add_item_to_order(order_id, photo, link_or_id, color, size, price, quantity, comment)

                excel_data = {
                    'Фото': photo,
                    'ID / Ссылка': link_or_id,
                    'Цвет': color,
                    'Размер': size,
                    'Цена за единицу': round(price * currency_rates.get(country.upper(), 1), 2),
                    'Кол-во/шт': quantity,
                    'tg контакт': telegram_username,
                    'Страна выкупа': country.upper(),
                    'комментарий': comment
                }

                save_to_excel(order_id, [item], full_name, telegram_username, country, "На модерации", "", "")
            
            # Уведомление менеджерам
            for manager_id in MANAGER_IDS:
                try:
                    bot.send_message(manager_id, f"🔔 Новый заказ #{order_id} от @{telegram_username}\nЖдёт подтверждения.", reply_markup=types.InlineKeyboardMarkup().add(
                        types.InlineKeyboardButton("Перейти в панель", callback_data="view_all_orders")
                    ))
                    log_event("manager_notify", user_id, f"Менеджеру {manager_id} отправлено уведомление")
                except Exception as e:
                    log_event("manager_notify_error", user_id, str(e))

            bot.send_message(message.chat.id, f"✅ Ваш заказ #{order_id} создан!\nОжидайте подтверждения менеджера.")

        elif data.get('command') == 'update_order_status':
            order_id = data.get('order_id')
            new_status = data.get('status')
            requisites = data.get('requisites', '')
            manager_comment = data.get('manager_comment', '')
            if order_id and new_status:
                update_order_status(order_id, new_status, requisites, manager_comment)
                bot.send_message(message.chat.id, f"🔄 Статус заказа #{order_id} изменён на '{new_status}'")

    except Exception as e:
        error_msg = str(e).replace("\n", "\\n")
        log_event("error", user_id, error_msg)
        bot.send_message(message.chat.id, f"❌ Ошибка при обработке данных: {error_msg}")

# --- Проверка регистрации ---
def is_registered(user_id):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT telegram_id FROM users WHERE telegram_id = ?", (user_id,))
    result = cursor.fetchone()
    conn.close()
    return bool(result)

# --- Команда /start ---
@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id
    full_name = message.from_user.full_name
    username = message.from_user.username or ""
    add_user(user_id, full_name, username)
    send_options(message)

# --- Главное меню ---
def send_options(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn_open_webapp = types.KeyboardButton("🧮КАЛЬКУЛЯТОР🧮")
    btn_my_orders = types.KeyboardButton("📦 Мои заказы")
    btn_manager = types.KeyboardButton("🔒 Панель менеджера")
    markup.add(btn_open_webapp, btn_my_orders, btn_manager)
    bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)

# --- Открытие WebApp ---
@bot.message_handler(func=lambda m: m.text in ["🧮КАЛЬКУЛЯТОР🧮", "📦 Мои заказы", "🔒 Панель менеджера"])
def open_webapp(message):
    webapp_url = "https://m37avolt.github.io/volt-logistics/" 
    markup = types.InlineKeyboardMarkup()
    btn_open = types.InlineKeyboardButton(text="🔗 Открыть WebApp", web_app=types.WebAppInfo(url=webapp_url))
    markup.add(btn_open)
    bot.send_message(message.chat.id, "🌐 Откройте WebApp:", reply_markup=markup)

# --- Является ли пользователь менеджером ---
def is_manager(user_id):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT is_manager FROM users WHERE telegram_id = ?", (user_id,))
    result = cursor.fetchone()
    conn.close()
    return bool(result[0]) if result else False

# --- Получение всех заказов ---
def get_all_orders():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT id, status, country FROM orders")
    result = cursor.fetchall()
    conn.close()
    return result

# --- Менеджерская панель ---
@bot.message_handler(func=lambda m: m.text == "🔒 Панель менеджера")
def manager_panel(message):
    if message.from_user.id in MANAGER_IDS:
        all_orders = get_all_orders()
        response = "📦 Все заказы:\n" + ("\n".join([
            f"• Заказ #{o[0]} | Статус: {o[4]} | Страна: {o[3]}"
            for o in all_orders
        ]) if all_orders else "Нет активных заказов.")
        bot.send_message(message.chat.id, response)

        if all_orders:
            markup = types.InlineKeyboardMarkup(row_width=1)
            for order in all_orders:
                markup.add(types.InlineKeyboardButton(f"Изменить статус #{order[0]}", callback_data=f"edit_status_{order[0]}"))
            bot.send_message(message.chat.id, "Выберите заказ для редактирования:", reply_markup=markup)
    else:
        bot.send_message(message.chat.id, "❌ Нет доступа к панели.")

# --- Callback handler ---
@bot.callback_query_handler(func=lambda call: True)
def handle_callback(call):
    if call.data.startswith("edit_status_"):
        order_id = int(call.data.split("_")[2])
        statuses = ["На модерации", "В работе", "Готов к оплате", "Оплачен", "Отправлен", "Завершён"]
        markup = types.InlineKeyboardMarkup(row_width=1)
        for status in statuses:
            markup.add(types.InlineKeyboardButton(status, callback_data=f"set_status_{order_id}_{status}"))
        bot.send_message(call.message.chat.id, f"Выберите новый статус для заказа #{order_id}:", reply_markup=markup)

    elif call.data.startswith("set_status_"):
        _, _, order_id, new_status = call.data.split("_", 3)
        order_id = int(order_id)
        update_order_status(order_id, new_status)
        bot.send_message(call.message.chat.id, f"🔄 Статус заказа #{order_id} изменён на '{new_status}'")

# --- Обновление статуса заказа ---
def update_order_status(order_id, new_status, requisites='', comment=''):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("UPDATE orders SET status=?, requisites=?, manager_comment=? WHERE id=?", 
                   (new_status, requisites, comment, order_id))
    conn.commit()
    conn.close()
    log_event("status_updated", "system", f"Заказ #{order_id} → {new_status}")

# --- Обработчик обычных сообщений ---
@bot.message_handler(func=lambda message: True)
def any_message_handler(message):
    send_options(message)

# --- Запуск бота ---
if __name__ == "__main__":
    print("🟢 Бот запущен...")
    bot.polling(none_stop=True)
